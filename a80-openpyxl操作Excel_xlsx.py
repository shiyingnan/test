# -*- coding:utf-8 -*-

import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter,column_index_from_string
from collections import Counter



def filter(filename,sheetname,condition):
    wb=openpyxl.load_workbook(filename)
    ws=wb[sheetname]

    wb.create_sheet('result')
    ws_res=wb['result']

    list=[]
    ll=[]
    t=len(condition)
    cvalue=tuple(condition.values())
    a = 1
    for row in ws.rows:
        if a==1:
            for cell in row:
                ll.append(cell.value)
            ws_res.append(ll)
            a+=1
            continue
        for cell in row:
            if str(cell.value) in cvalue[0]:    #筛选条件1
                list.append(row)
    temp,list=list,[]

    for time in range(t-1):
        for tp in temp:
            for tmpcell in tp:
                if str(tmpcell.value) in cvalue[time+1]:   #筛选条件2
                    list.append(tp)
        temp,list=list,[]
    return temp


def excf(filename,sheetname,condition):
    '''
    筛选后结果写入result，同一个excel的sheet表内
    :param filename: excel文件地址
    :param sheetname: 要操作的sheet
    :param condition: 筛选条件
    :return:
    '''
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    sheetnames = wb.sheetnames
    print(sheetnames)

    for name in sheetnames:
        if name == 'result':
            del wb['result']

    wb.create_sheet('result')
    ws_res = wb['result']

    slist = []
    frow=[]

    citems = list(condition.items())
    print(citems)
    for r in ws['1']:
        frow.append(r.value)
    ws_res.append(frow)

    for colcell in ws[citems[0][0]]:
        if str(colcell.value) in citems[0][1]:   # 筛选条件1
            row=str(colcell.row)
            slist.append(ws[row])
    temp,slist=slist,[]
    del citems[0]
    print(citems)

    t = len(citems)
    for time in range(t):
        for col in temp:
            colnum=column_index_from_string(citems[time][0])
            if str(col[colnum-1].value) in citems[time][1]:
                slist.append(col)
        temp,slist=slist,[]

    for w in temp:
        wlist = []
        for write in w:
            wlist.append(write.value)
        ws_res.append(wlist)

    wb.save('file.xlsx')

def tongji(filename,sheetname,column_str,column_sum=False):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    sheetnames = wb.sheetnames
    print(sheetnames)
    print('共有行数:',ws.max_row)   #所有工单数
    #统计无重复的工单数
    e_list=[]
    for i in ws[column_str]:
        e_list.append(str(i.value))
    e_list.remove(e_list[0])       #删除‘qq’表头
    e_set=set(e_list)
    e_lenth=len(e_set)
    print('%s列去重后数据共有:%d个'%(ws[column_str+'1'].value,e_lenth)) #统计无重复的工单数

    e_counter=Counter(e_list)
    e_dict=dict(e_counter)
    for j in e_dict:
        if e_dict[j]>1:
            print('重复数据：%s，个数为：%d'%(j,e_dict[j]))   #显示重复的数据及个数

#统计数据之和
    if column_sum==True:
        sum_list=[]
        for k in e_list:
            if k=='None':
                continue
            else:
                m=int(k)
                sum_list.append(m)
        zonghe=sum(sum_list)
        print('%s列统计之和为：%d'%(ws[column_str+'1'].value,zonghe))



if __name__=='__main__':
    con = {"B": "13级",
           "C": "采矿工程",
           "F": "中国矿业大学"

            }
    excf('file.xlsx','syntest',con)
    tongji('file.xlsx','result','G',True)








    # for ii in temp:
    #     listout = []
    #     for jcell in ii:
    #         listout.append(jcell.value)
    #     ws_res.append(listout)
    #
    # wb.save('file.xlsx')











# def readwb(wbname,sheetname):
#     wb=openpyxl.load_workbook(filename=wbname,read_only=True)
#     if (sheetname==""):
#         ws=wb.active
#     else:
#         ws=wb[sheetname]
#     i=1
#     fields=[]
#     data=[]
#     for row in ws.rows:
#         list=[]
#         for cell in row:
#             aa=str(cell.value)
#             if (aa=='None'):
#                 aa='1'
#             list.append(aa)
#         if (i==1):
#             fields=list
#         else:
#             data.append(list)
#         i+=1
#     # data.sort(key=lambda x: x[0])
#     print(wbname+"-"+sheetname+"-已成功读取")
#     return fields,data
#
# print(readwb(("file.xlsx"),""))





# def filt(wbname,sheetname,condition,outwb):
#     fields=[]
#     data=[]
#     tmp=[]
#     tmplst=[]
#     fields,data=readwb(wbname,sheetname)
#     count=0
#     confirm=0
#     for field,value in condition.items():
#         if count==0:
#             tmplst=copy.deepcopy(data)
#         else:
#             tmplst=copy.deepcopy(tmp)
#         try:
#             idx=fields.index(field)
#         except:
#             print(field+'not exists')
#             continue
#         for iterm in tmplst:
#             confirm=0
#             for val in value:
#                 if (val in iterm[idx]) and (count==0) and (iterm not in tmp):
#                     tmp.append(iterm)
#                 if (count!=0) and (val not in iterm[idx]):
#                     confirm+=1
#                     if confirm == len(value):
#                         tmp.remove(iterm)
#         count +=1













# wb=openpyxl.load_workbook("file.xlsx")   #加载excel文件
#
# ws=wb['syntest']
#
#
# ws.auto_filter.ref = "A1:E12"
# ws.auto_filter.add_filter_column(0, ["11级", "12级"])
# ws.auto_filter.add_sort_condition("B2:B12")
#
#
#
#
# wb.save("filtered.xlsx")


#-----------------------读取第1行的所有值-----------------------------
# print(sheet['1'])
# for i in sheet['1']:
#     print(i.value)
#------------------------读取第2列的所有值----------------------------

# fpy=open('tes.py','w',encoding='utf-8')
#
# print(sheet['B'])
# for j in sheet['B']:
#     print(j.value)
#     fpy.write(str(j.value)+'\n')
# fpy.close()
# wb = Workbook()
# ws = wb.active
#
# a1 = ws['A1']
# # a1.value=4
# #
# # ws['B1'].value=87
# # ws["C1"] = "=SUM(A1, B1)"       #求和公式
#
# # italic24Font = Font(color=colors.RED)
# #
# # a1.font=italic24Font
#
# wb.save('yangshi.xlsx')


# t=tuple(sheet['A1':'C3'])
# for i in t:
#     for j in i:
#         print(j.value)





# sheetnames=wb.sheetnames           #查看sheet
# print(sheetnames)

# ws=wb["2027测试表"]
#
# for i in ws["2"]:
#     print(i.value,end="\t")
# print()
#
# print("第2行第5列的值:",ws["E2"].value)  #输出第2行第5列的值
#
# print(ws.cell(row=2,column=5).value)   #输出第2行第5列的值
#
# print(ws.max_column)       #sheet表的列数
# print(ws.max_row)         #sheet表的行数
# rlist=[]
# print(tuple(ws["A1":"C3"]))
#
# for j in ws["A2":"F2"]:       #取出第二行的所有数据，空单元格为None
#     for k in j:
#         rlist.append(k.value)
# print(tuple(rlist))
#
# clist=[]
# for j in ws["B1":"B4"]:       #取出第二列的所有数据
#     for k in j:
#         clist.append(k.value)
# print(tuple(clist))

